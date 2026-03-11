"""
Reads Microsoft Forms responses from an Excel file (Form Responses 1),
then writes each response into specific columns of a destination Excel table.

Requires:
  pip install openpyxl


from datetime import datetime
from openpyxl import load_workbook

# --- CONFIG ---
SOURCE_XLSX = r"./FormResponses.xlsx"      # the workbook that contains the form response sheet
SOURCE_SHEET = "Form Responses 1"          # common default name
DEST_XLSX = r"./MasterTracker.xlsx"        # your real workbook
DEST_SHEET = "Tracker"                     # sheet with your table

# A mapping from "source column header" -> "destination column header"
# Edit these to match your real column names.
COLUMN_MAP = {
    "Timestamp": "Submitted At",
    "Name": "Employee Name",
    "Email": "Email",
    "Issue Type": "Category",
    "Description": "Details",
    "Site": "Location",
}

# Optional: where to store a "last imported timestamp" so you don't import duplicates.
# You can also do this by storing a hash/ID.
STATE_CELL = "Z1"  # in DEST_SHEET


def header_index_map(ws, header_row=1):
    """Return dict: header text -> column index (1-based)."""
    m = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            m[str(cell.value).strip()] = cell.col_idx
    return m


def find_next_empty_row(ws, required_col=1, start_row=2):
    """Find next empty row by scanning a 'required' column (default col A)."""
    r = start_row
    while ws.cell(row=r, column=required_col).value not in (None, ""):
        r += 1
    return r


def parse_timestamp(value):
    """Try to parse a Forms timestamp cell into a datetime."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        # Forms exports often look like: "3/3/2026 2:15:30 PM"
        for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                pass
    return None


def main():
    # Load source responses workbook
    src_wb = load_workbook(SOURCE_XLSX, data_only=True)
    src_ws = src_wb[SOURCE_SHEET]
    src_headers = header_index_map(src_ws, header_row=1)

    # Load destination workbook
    dst_wb = load_workbook(DEST_XLSX)
    dst_ws = dst_wb[DEST_SHEET]
    dst_headers = header_index_map(dst_ws, header_row=1)

    # Figure out last imported timestamp (optional)
    last_ts = dst_ws[STATE_CELL].value
    last_ts_dt = parse_timestamp(last_ts) if last_ts else None

    # Identify which column in source holds timestamp (if you want incremental import)
    ts_col = src_headers.get("Timestamp")

    imported = 0

    # Iterate through source rows (starting after header)
    for row in range(2, src_ws.max_row + 1):
        # If using incremental import, skip anything older/equal than last imported timestamp
        if ts_col is not None:
            row_ts = parse_timestamp(src_ws.cell(row=row, column=ts_col).value)
            if last_ts_dt and row_ts and row_ts <= last_ts_dt:
                continue

        # Determine destination row to write to
        dest_row = find_next_empty_row(dst_ws, required_col=1, start_row=2)

        # Copy mapped fields
        for src_header, dst_header in COLUMN_MAP.items():
            if src_header not in src_headers:
                raise KeyError(f"Source header not found: {src_header}")
            if dst_header not in dst_headers:
                raise KeyError(f"Destination header not found: {dst_header}")

            value = src_ws.cell(row=row, column=src_headers[src_header]).value
            dst_ws.cell(row=dest_row, column=dst_headers[dst_header]).value = value

        imported += 1

    # Update state (store newest timestamp seen)
    if imported > 0 and ts_col is not None:
        # find max timestamp in source (simple approach)
        newest = None
        for row in range(2, src_ws.max_row + 1):
            t = parse_timestamp(src_ws.cell(row=row, column=ts_col).value)
            if t and (newest is None or t > newest):
                newest = t
        if newest:
            dst_ws[STATE_CELL].value = newest

    dst_wb.save(DEST_XLSX)
    print(f"Imported {imported} new form response(s).")


if __name__ == "__main__":
    main()
    """

"""all commented for now i guess"""